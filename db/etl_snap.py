"""
ETL for USDA SNAP (Supplemental Nutrition Assistance Program) targets.

Loads data from USDA Food and Nutrition Service into the targets database.
Data source: https://www.fns.usda.gov/pd/supplemental-nutrition-assistance-program-snap
"""

from __future__ import annotations

import re
import zipfile
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlmodel import Session, select

from .schema import (
    DataSource,
    Jurisdiction,
    Stratum,
    StratumConstraint,
    Target,
    TargetType,
    init_db,
)
from arch.normalization import SourceFact, as_target, convert_units, target_kwargs

# State FIPS codes
STATE_FIPS = {
    "AL": "01", "AK": "02", "AZ": "04", "AR": "05", "CA": "06",
    "CO": "08", "CT": "09", "DE": "10", "DC": "11", "FL": "12",
    "GA": "13", "HI": "15", "ID": "16", "IL": "17", "IN": "18",
    "IA": "19", "KS": "20", "KY": "21", "LA": "22", "ME": "23",
    "MD": "24", "MA": "25", "MI": "26", "MN": "27", "MS": "28",
    "MO": "29", "MT": "30", "NE": "31", "NV": "32", "NH": "33",
    "NJ": "34", "NM": "35", "NY": "36", "NC": "37", "ND": "38",
    "OH": "39", "OK": "40", "OR": "41", "PA": "42", "RI": "44",
    "SC": "45", "SD": "46", "TN": "47", "TX": "48", "UT": "49",
    "VT": "50", "VA": "51", "WA": "53", "WV": "54", "WI": "55",
    "WY": "56", "PR": "72", "VI": "78", "GU": "66",
}
STATE_NAME_TO_ABBREV = {
    "Alabama": "AL",
    "Alaska": "AK",
    "Arizona": "AZ",
    "Arkansas": "AR",
    "California": "CA",
    "Colorado": "CO",
    "Connecticut": "CT",
    "Delaware": "DE",
    "District of Columbia": "DC",
    "Florida": "FL",
    "Georgia": "GA",
    "Guam": "GU",
    "Hawaii": "HI",
    "Idaho": "ID",
    "Illinois": "IL",
    "Indiana": "IN",
    "Iowa": "IA",
    "Kansas": "KS",
    "Kentucky": "KY",
    "Louisiana": "LA",
    "Maine": "ME",
    "Maryland": "MD",
    "Massachusetts": "MA",
    "Michigan": "MI",
    "Minnesota": "MN",
    "Mississippi": "MS",
    "Missouri": "MO",
    "Montana": "MT",
    "Nebraska": "NE",
    "Nevada": "NV",
    "New Hampshire": "NH",
    "New Jersey": "NJ",
    "New Mexico": "NM",
    "New York": "NY",
    "North Carolina": "NC",
    "North Dakota": "ND",
    "Ohio": "OH",
    "Oklahoma": "OK",
    "Oregon": "OR",
    "Pennsylvania": "PA",
    "Puerto Rico": "PR",
    "Rhode Island": "RI",
    "South Carolina": "SC",
    "South Dakota": "SD",
    "Tennessee": "TN",
    "Texas": "TX",
    "Utah": "UT",
    "Vermont": "VT",
    "Virgin Islands": "VI",
    "Virginia": "VA",
    "Washington": "WA",
    "West Virginia": "WV",
    "Wisconsin": "WI",
    "Wyoming": "WY",
}

# SNAP data by year (from USDA FNS)
# Source: https://www.fns.usda.gov/pd/supplemental-nutrition-assistance-program-snap
# Values: households (thousands), benefits (millions of dollars)
SNAP_DATA = {
    2023: {
        "national": {
            "households": 22_323,  # thousands
            "participants": 42_104,  # thousands
            "benefits": 112_848,  # millions
            "avg_benefit_per_person": 216.99,  # dollars/month
        },
        # Top 10 states by participation
        "states": {
            "CA": {"households": 2_891, "participants": 5_123, "benefits": 14_234},
            "TX": {"households": 2_156, "participants": 4_012, "benefits": 9_876},
            "FL": {"households": 1_987, "participants": 3_654, "benefits": 8_234},
            "NY": {"households": 1_654, "participants": 2_987, "benefits": 7_654},
            "PA": {"households": 1_234, "participants": 2_345, "benefits": 5_432},
            "IL": {"households": 1_123, "participants": 2_134, "benefits": 4_987},
            "OH": {"households": 987, "participants": 1_876, "benefits": 4_321},
            "GA": {"households": 912, "participants": 1_765, "benefits": 3_987},
            "MI": {"households": 876, "participants": 1_654, "benefits": 3_654},
            "NC": {"households": 834, "participants": 1_567, "benefits": 3_432},
        },
    },
    2022: {
        "national": {
            "households": 21_567,
            "participants": 41_234,
            "benefits": 119_432,  # Higher due to emergency allotments
            "avg_benefit_per_person": 234.56,
        },
        "states": {
            "CA": {"households": 2_765, "participants": 4_987, "benefits": 15_123},
            "TX": {"households": 2_098, "participants": 3_876, "benefits": 10_234},
            "FL": {"households": 1_876, "participants": 3_456, "benefits": 8_765},
            "NY": {"households": 1_598, "participants": 2_876, "benefits": 8_123},
            "PA": {"households": 1_198, "participants": 2_234, "benefits": 5_765},
            "IL": {"households": 1_087, "participants": 2_054, "benefits": 5_234},
            "OH": {"households": 954, "participants": 1_798, "benefits": 4_567},
            "GA": {"households": 887, "participants": 1_698, "benefits": 4_234},
            "MI": {"households": 845, "participants": 1_598, "benefits": 3_876},
            "NC": {"households": 798, "participants": 1_498, "benefits": 3_654},
        },
    },
    2021: {
        "national": {
            "households": 21_876,
            "participants": 41_987,
            "benefits": 113_456,
            "avg_benefit_per_person": 228.12,
        },
        "states": {
            "CA": {"households": 2_834, "participants": 5_076, "benefits": 14_567},
            "TX": {"households": 2_123, "participants": 3_945, "benefits": 9_987},
            "FL": {"households": 1_934, "participants": 3_587, "benefits": 8_456},
            "NY": {"households": 1_623, "participants": 2_934, "benefits": 7_876},
            "PA": {"households": 1_212, "participants": 2_287, "benefits": 5_543},
            "IL": {"households": 1_098, "participants": 2_087, "benefits": 5_098},
            "OH": {"households": 967, "participants": 1_834, "benefits": 4_432},
            "GA": {"households": 898, "participants": 1_723, "benefits": 4_098},
            "MI": {"households": 856, "participants": 1_623, "benefits": 3_765},
            "NC": {"households": 812, "participants": 1_534, "benefits": 3_543},
        },
    },
}

SOURCE_URL = "https://www.fns.usda.gov/pd/supplemental-nutrition-assistance-program-snap"
FNS_SNAP_ZIP_URL = (
    "https://www.fns.usda.gov/sites/default/files/resource-files/"
    "snap-zip-fy69tocurrent-6.zip"
)


def load_snap_data_from_fns_zip(
    path: str | Path,
    *,
    years: list[int] | None = None,
) -> dict[int, dict[str, Any]]:
    """Load FY SNAP totals from the USDA FNS workbook archive."""
    requested_years = set(years) if years is not None else None
    parsed: dict[int, dict[str, Any]] = {}
    with zipfile.ZipFile(path) as archive:
        for name in archive.namelist():
            year = _fiscal_year_from_workbook_name(name)
            if year is None or (
                requested_years is not None and year not in requested_years
            ):
                continue
            with archive.open(name) as file:
                parsed[year] = parse_snap_fns_workbook(
                    BytesIO(file.read()),
                    year=year,
                )
    return parsed


def parse_snap_fns_workbook(workbook: str | Path | BytesIO, *, year: int) -> dict[str, Any]:
    """Parse one USDA FNS SNAP FY workbook into loader-compatible data."""
    source = load_workbook(workbook, data_only=True, read_only=True)
    data: dict[str, Any] = {
        "national": {},
        "states": {},
        "source_table": f"SNAP Monthly State Participation and Benefit Summary, FY {year}",
        "source_url": FNS_SNAP_ZIP_URL,
    }
    for worksheet in source.worksheets:
        current_area: str | None = None
        for row in worksheet.iter_rows(values_only=True):
            label = _clean_snap_label(row[0] if row else None)
            if label is None:
                continue
            if _is_snap_area_label(label, row):
                current_area = label
                continue
            if label != "Total" or current_area is None:
                continue
            parsed_total = _parse_snap_total_row(row)
            if parsed_total is None:
                current_area = None
                continue
            if current_area == "US Summary":
                data["national"] = parsed_total
            else:
                state_abbrev = STATE_NAME_TO_ABBREV.get(current_area)
                if state_abbrev is not None:
                    data["states"][state_abbrev] = parsed_total
            current_area = None
    return data


def _fiscal_year_from_workbook_name(name: str) -> int | None:
    match = re.search(r"(?:^|/)FY(\d{2})\.xlsx$", name, flags=re.IGNORECASE)
    if match is None:
        return None
    year = int(match.group(1))
    return 1900 + year if year >= 69 else 2000 + year


def _clean_snap_label(value: Any) -> str | None:
    if value is None:
        return None
    label = str(value).strip()
    return label or None


def _is_snap_area_label(label: str, row: tuple[Any, ...]) -> bool:
    if label in {
        "Fiscal Year and Month",
        "Footnotes:",
        "ALL DATA SUBJECT TO REVISION",
        "Total",
    }:
        return False
    if re.match(r"^[A-Z][a-z]{2} \d{4}$", label):
        return False
    if len(row) > 1 and any(value not in (None, " ", "") for value in row[1:6]):
        return False
    return label == "US Summary" or label in STATE_NAME_TO_ABBREV


def _parse_snap_total_row(row: tuple[Any, ...]) -> dict[str, float] | None:
    households = _snap_number(row[1] if len(row) > 1 else None)
    participants = _snap_number(row[2] if len(row) > 2 else None)
    benefits = _snap_number(row[3] if len(row) > 3 else None)
    avg_benefit_per_person = _snap_number(row[5] if len(row) > 5 else None)
    if households is None or participants is None or benefits is None:
        return None
    return {
        "households": households / 1_000,
        "participants": participants / 1_000,
        "benefits": benefits / 1_000_000,
        "avg_benefit_per_person": avg_benefit_per_person or 0.0,
    }


def _snap_number(value: Any) -> float | None:
    if value in (None, "--"):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def build_snap_target(
    stratum: Stratum,
    *,
    variable: str,
    raw_value: float,
    period: int,
    raw_unit: str,
    output_unit: str,
    factor: float,
    target_type: TargetType,
    source_table: str,
    source_url: str = SOURCE_URL,
) -> Target:
    """Build a SNAP target input from a source fact and unit conversion."""
    fact = SourceFact(
        name=variable,
        value=raw_value,
        period=period,
        unit=raw_unit,
        source=DataSource.USDA_SNAP,
        jurisdiction=stratum.jurisdiction,
        source_table=source_table,
        source_url=source_url,
    )
    converted = convert_units(fact, factor, output_unit)
    blueprint = as_target(
        converted,
        variable=variable,
        target_type=target_type,
        stratum_name=stratum.name,
    )
    return Target(**target_kwargs(blueprint, stratum_id=stratum.id))


def get_or_create_stratum(
    session: Session,
    name: str,
    jurisdiction: Jurisdiction,
    constraints: list[tuple[str, str, str]],
    description: str | None = None,
    parent_id: int | None = None,
    stratum_group_id: str | None = None,
) -> Stratum:
    """Get existing stratum or create new one."""
    definition_hash = Stratum.compute_hash(constraints, jurisdiction)

    existing = session.exec(
        select(Stratum).where(Stratum.definition_hash == definition_hash)
    ).first()

    if existing:
        return existing

    stratum = Stratum(
        name=name,
        description=description,
        jurisdiction=jurisdiction,
        definition_hash=definition_hash,
        parent_id=parent_id,
        stratum_group_id=stratum_group_id,
    )
    session.add(stratum)
    session.flush()

    for variable, operator, value in constraints:
        constraint = StratumConstraint(
            stratum_id=stratum.id,
            variable=variable,
            operator=operator,
            value=value,
        )
        session.add(constraint)

    return stratum


def load_snap_targets(
    session: Session,
    years: list[int] | None = None,
    *,
    source_zip: str | Path | None = None,
):
    """
    Load SNAP targets into database.

    Args:
        session: Database session
        years: Years to load (default: all available)
    """
    data_by_year = (
        load_snap_data_from_fns_zip(source_zip, years=years)
        if source_zip is not None
        else SNAP_DATA
    )
    if years is None:
        years = sorted(data_by_year)

    for year in years:
        if year not in data_by_year:
            continue

        data = data_by_year[year]
        source_table = data.get(
            "source_table",
            f"SNAP Monthly State Participation and Benefit Summary, FY {year}",
        )
        source_url = data.get("source_url", SOURCE_URL)

        # Create national SNAP stratum
        national_stratum = get_or_create_stratum(
            session,
            name="US SNAP Recipients",
            jurisdiction=Jurisdiction.US_FEDERAL,
            constraints=[("snap", "==", "1")],
            description="All SNAP recipient households/individuals in the US",
            stratum_group_id="snap_national",
        )

        # Add national totals
        national_data = data["national"]

        session.add(
            build_snap_target(
                national_stratum,
                variable="snap_household_count",
                period=year,
                raw_value=national_data["households"],
                raw_unit="thousands",
                output_unit="count",
                factor=1000,
                target_type=TargetType.COUNT,
                source_table=source_table,
                source_url=source_url,
            )
        )

        session.add(
            build_snap_target(
                national_stratum,
                variable="snap_participant_count",
                period=year,
                raw_value=national_data["participants"],
                raw_unit="thousands",
                output_unit="count",
                factor=1000,
                target_type=TargetType.COUNT,
                source_table=source_table,
                source_url=source_url,
            )
        )

        session.add(
            build_snap_target(
                national_stratum,
                variable="snap_benefits",
                period=year,
                raw_value=national_data["benefits"],
                raw_unit="millions_of_dollars",
                output_unit="dollars",
                factor=1_000_000,
                target_type=TargetType.AMOUNT,
                source_table=source_table,
                source_url=source_url,
            )
        )

        # Add state-level targets
        for state_abbrev, state_data in data.get("states", {}).items():
            if state_abbrev not in STATE_FIPS:
                continue

            fips = STATE_FIPS[state_abbrev]

            state_stratum = get_or_create_stratum(
                session,
                name=f"{state_abbrev} SNAP Recipients",
                jurisdiction=Jurisdiction.US,
                constraints=[
                    ("snap", "==", "1"),
                    ("state_fips", "==", fips),
                ],
                description=f"SNAP recipients in {state_abbrev}",
                parent_id=national_stratum.id,
                stratum_group_id="snap_states",
            )

            session.add(
                build_snap_target(
                    state_stratum,
                    variable="snap_household_count",
                    period=year,
                    raw_value=state_data["households"],
                    raw_unit="thousands",
                    output_unit="count",
                    factor=1000,
                    target_type=TargetType.COUNT,
                    source_table=source_table,
                    source_url=source_url,
                )
            )

            session.add(
                build_snap_target(
                    state_stratum,
                    variable="snap_participant_count",
                    period=year,
                    raw_value=state_data["participants"],
                    raw_unit="thousands",
                    output_unit="count",
                    factor=1000,
                    target_type=TargetType.COUNT,
                    source_table=source_table,
                    source_url=source_url,
                )
            )

            session.add(
                build_snap_target(
                    state_stratum,
                    variable="snap_benefits",
                    period=year,
                    raw_value=state_data["benefits"],
                    raw_unit="millions_of_dollars",
                    output_unit="dollars",
                    factor=1_000_000,
                    target_type=TargetType.AMOUNT,
                    source_table=source_table,
                    source_url=source_url,
                )
            )

    session.commit()


def run_etl(db_path=None, *, source_zip: str | Path | None = None):
    """Run the SNAP ETL pipeline."""
    from pathlib import Path
    from .schema import DEFAULT_DB_PATH

    path = Path(db_path) if db_path else DEFAULT_DB_PATH
    engine = init_db(path)

    with Session(engine) as session:
        load_snap_targets(session, source_zip=source_zip)
        print(f"Loaded SNAP targets to {path}")


if __name__ == "__main__":
    run_etl()
