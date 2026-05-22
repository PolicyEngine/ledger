"""ETL for HHS ACF TANF financial targets."""

from __future__ import annotations

import hashlib
from functools import lru_cache
from importlib.resources import files
from io import BytesIO
from typing import Any, TypedDict

import yaml
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import delete
from sqlmodel import Session

from .etl_snap import STATE_NAME_TO_ABBREV
from .etl_soi import get_or_create_stratum
from .etl_soi_state import STATE_FIPS
from .schema import DataSource, GeographicLevel, Jurisdiction, Target, TargetType

PACKAGE_DIR = "data/hhs_acf/tanf_financial_2024"
MANIFEST = "manifest.yaml"
SOURCE_TABLE = "FY 2024 Federal TANF and State MOE Financial Data"
NATIONAL_SHEET = "A.1 Fed & State by Category"
VARIABLE = "tanf_cash_assistance"
NARROW_BASIC_ASSISTANCE_ROW = (
    "Basic Assistance (excluding Relative Foster Care Maintenance Payments "
    "and Adoption and Guardianship Subsidies)"
)


class TANFFinancialData(TypedDict):
    """Parsed TANF cash-assistance facts from an ACF financial workbook."""

    source_url: str
    national_cash_assistance: float
    states: dict[str, float]


@lru_cache(maxsize=1)
def _manifest() -> dict[str, Any]:
    manifest_path = files("db").joinpath(PACKAGE_DIR, MANIFEST)
    with manifest_path.open("r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def available_tanf_financial_years() -> list[int]:
    """Return packaged ACF TANF financial workbook years."""
    return sorted(int(year) for year in _manifest()["files"])


def _file_spec(year: int) -> dict[str, Any]:
    files_by_year = _manifest()["files"]
    try:
        return files_by_year[year]
    except KeyError:
        return files_by_year[str(year)]


def tanf_financial_source_url(year: int) -> str:
    """Return the ACF source page URL for a TANF financial workbook year."""
    return str(_file_spec(year)["source_url"])


@lru_cache(maxsize=None)
def _content(year: int) -> bytes:
    spec = _file_spec(year)
    package_path = files("db").joinpath(PACKAGE_DIR, spec["filename"])
    content = package_path.read_bytes()
    expected_sha = spec.get("sha256")
    if expected_sha:
        actual_sha = hashlib.sha256(content).hexdigest()
        if actual_sha != expected_sha:
            raise ValueError(
                f"ACF TANF financial workbook {year} checksum mismatch: "
                f"expected {expected_sha}, got {actual_sha}"
            )
    return content


@lru_cache(maxsize=None)
def load_tanf_financial_data(year: int) -> TANFFinancialData:
    """Parse national and state TANF cash assistance from ACF financial data."""
    workbook = load_workbook(
        filename=BytesIO(_content(year)),
        read_only=True,
        data_only=True,
    )
    if NATIONAL_SHEET not in workbook.sheetnames:
        raise ValueError(f"ACF TANF financial workbook {year} missing {NATIONAL_SHEET}")

    states: dict[str, float] = {}
    for state_name, state_abbrev in STATE_NAME_TO_ABBREV.items():
        if state_abbrev not in STATE_FIPS:
            continue
        sheet_name = "DC" if state_abbrev == "DC" else state_name
        if sheet_name not in workbook.sheetnames:
            continue
        states[state_abbrev] = _extract_cash_assistance_all_funds(
            workbook[sheet_name]
        )

    expected_states = set(STATE_FIPS)
    if set(states) != expected_states:
        missing = ", ".join(sorted(expected_states - set(states)))
        extra = ", ".join(sorted(set(states) - expected_states))
        raise ValueError(
            f"ACF TANF financial workbook {year} state coverage mismatch: "
            f"missing={missing or 'none'}; extra={extra or 'none'}"
        )

    return {
        "source_url": tanf_financial_source_url(year),
        "national_cash_assistance": _extract_cash_assistance_all_funds(
            workbook[NATIONAL_SHEET]
        ),
        "states": states,
    }


def load_tanf_targets(
    session: Session,
    years: list[int] | None = None,
) -> None:
    """Load ACF TANF cash-assistance targets into the database."""
    if years is None:
        years = available_tanf_financial_years()

    session.exec(
        delete(Target).where(
            Target.source == DataSource.HHS_ACF_TANF,
            Target.source_table == SOURCE_TABLE,
            Target.period.in_(years),
            Target.variable == VARIABLE,
        )
    )

    national_stratum = get_or_create_stratum(
        session,
        name="US population",
        jurisdiction=Jurisdiction.US,
        constraints=[],
        description="United States population",
        stratum_group_id="hhs_acf_tanf_national",
    )
    for year in years:
        if year not in available_tanf_financial_years():
            continue
        data = load_tanf_financial_data(year)
        _add_target(
            session,
            stratum_id=int(national_stratum.id),
            period=year,
            value=data["national_cash_assistance"],
            geographic_level=GeographicLevel.NATIONAL,
            source_url=data["source_url"],
        )
        for state_abbrev, value in data["states"].items():
            state_stratum = get_or_create_stratum(
                session,
                name=f"{state_abbrev} population",
                jurisdiction=Jurisdiction.US,
                constraints=[("state_fips", "==", STATE_FIPS[state_abbrev])],
                description=f"Population in {state_abbrev}",
                stratum_group_id="hhs_acf_tanf_states",
            )
            _add_target(
                session,
                stratum_id=int(state_stratum.id),
                period=year,
                value=value,
                geographic_level=GeographicLevel.STATE,
                source_url=data["source_url"],
            )
    session.commit()


def _add_target(
    session: Session,
    *,
    stratum_id: int,
    period: int,
    value: float,
    geographic_level: GeographicLevel,
    source_url: str,
) -> None:
    session.add(
        Target(
            stratum_id=stratum_id,
            variable=VARIABLE,
            period=period,
            value=value,
            target_type=TargetType.AMOUNT,
            geographic_level=geographic_level,
            source=DataSource.HHS_ACF_TANF,
            source_table=SOURCE_TABLE,
            source_url=source_url,
            notes=(
                "Basic Assistance excluding relative foster care maintenance "
                "payments and adoption and guardianship subsidies."
            ),
        )
    )


def _extract_cash_assistance_all_funds(sheet: Worksheet) -> float:
    header_row: tuple[Any, ...] | None = None
    for row in sheet.iter_rows(values_only=True):
        normalized = tuple(_normalized_header(value) for value in row)
        if "Spending Category" in normalized and "All Funds" in normalized:
            header_row = normalized
            break
    if header_row is None:
        raise ValueError(f"Unexpected ACF TANF sheet columns: {sheet.title}")

    category_index = header_row.index("Spending Category")
    all_funds_index = header_row.index("All Funds")
    for row in sheet.iter_rows(values_only=True):
        if len(row) <= max(category_index, all_funds_index):
            continue
        if str(row[category_index]).strip() == NARROW_BASIC_ASSISTANCE_ROW:
            return _money_value(row[all_funds_index])
    raise ValueError(
        f"Could not locate narrow Basic Assistance row in ACF TANF sheet {sheet.title}"
    )


def _normalized_header(value: object) -> str:
    return " ".join(str(value or "").split())


def _money_value(value: object) -> float:
    if value is None:
        raise ValueError("Expected ACF TANF money cell, got empty value")
    text = str(value).replace("$", "").replace(",", "").strip()
    if not text:
        raise ValueError("Expected ACF TANF money cell, got empty value")
    return round(float(text), 2)
