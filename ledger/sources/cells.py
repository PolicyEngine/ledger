"""Source-cell records for preserving parsed source artifacts."""

from __future__ import annotations

import csv
from collections import Counter
from dataclasses import asdict, dataclass
from datetime import date, datetime, time
import hashlib
from html.parser import HTMLParser
from io import BytesIO, StringIO
import json
from pathlib import Path
import re
from typing import Any
from xml.etree import ElementTree
from zipfile import ZipFile

import openpyxl
import xlrd

Scalar = str | int | float | bool | None
SOURCE_CELL_KEY_PREFIX = "ledger.source_cell.v1"


@dataclass(frozen=True)
class SourceArtifactMetadata:
    """Immutable source artifact identity for parsed cells."""

    source_name: str
    source_table: str
    source_file: str
    url: str | None
    vintage: str
    sha256: str
    size_bytes: int
    extracted_at: str
    extraction_method: str
    raw_r2_bucket: str | None = None
    raw_r2_key: str | None = None
    raw_r2_uri: str | None = None


@dataclass(frozen=True)
class SourceCell:
    """One parsed cell from a source artifact."""

    artifact: SourceArtifactMetadata
    sheet_name: str
    row_number: int
    column_number: int
    address: str
    cell_type: str
    raw_value: Scalar
    display_value: str | None
    formula: str | None = None
    note: str | None = None
    source_row_key: str | None = None


@dataclass(frozen=True)
class SourceCellIssue:
    """One source-cell validation issue."""

    code: str
    message: str
    source_cell_key: str | None = None
    cell_index: int | None = None

    def to_dict(self) -> dict[str, Any]:
        """Return a JSON-serializable issue."""
        data = asdict(self)
        return {key: value for key, value in data.items() if value is not None}


@dataclass(frozen=True)
class SourceCellReport:
    """Validation and QA summary for source-cell records."""

    cell_count: int
    counts: dict[str, dict[str, int]]
    errors: tuple[SourceCellIssue, ...]
    warnings: tuple[SourceCellIssue, ...] = ()

    @property
    def valid(self) -> bool:
        """Whether the cell set has no validation errors."""
        return not self.errors

    def to_dict(self) -> dict[str, Any]:
        """Return a JSON-serializable report."""
        return {
            "valid": self.valid,
            "cell_count": self.cell_count,
            "counts": self.counts,
            "errors": [issue.to_dict() for issue in self.errors],
            "warnings": [issue.to_dict() for issue in self.warnings],
        }


def build_source_cell_key(cell: SourceCell) -> str:
    """Build a stable key from artifact hash and sheet coordinates."""
    payload = {
        "artifact_sha256": cell.artifact.sha256,
        "sheet_name": cell.sheet_name,
        "row_number": cell.row_number,
        "column_number": cell.column_number,
    }
    raw = json.dumps(payload, sort_keys=True, separators=(",", ":"))
    digest = hashlib.sha256(raw.encode("utf-8")).hexdigest()[:24]
    return f"{SOURCE_CELL_KEY_PREFIX}:{digest}"


def source_cells_from_xls(
    content: bytes,
    artifact: SourceArtifactMetadata,
) -> list[SourceCell]:
    """Parse all used-range cells from an XLS workbook."""
    workbook = xlrd.open_workbook(file_contents=content, formatting_info=False)
    cells: list[SourceCell] = []
    for sheet in workbook.sheets():
        for row_index in range(sheet.nrows):
            for column_index in range(sheet.ncols):
                cell = sheet.cell(row_index, column_index)
                raw_value = _cell_raw_value(cell, workbook.datemode)
                display_value = None if raw_value is None else str(raw_value)
                cells.append(
                    SourceCell(
                        artifact=artifact,
                        sheet_name=sheet.name,
                        row_number=row_index + 1,
                        column_number=column_index + 1,
                        address=f"{_excel_column_name(column_index + 1)}{row_index + 1}",
                        cell_type=_cell_type_name(cell.ctype),
                        raw_value=raw_value,
                        display_value=display_value,
                    )
                )
    return cells


def source_cells_from_xlsx(
    content: bytes,
    artifact: SourceArtifactMetadata,
) -> list[SourceCell]:
    """Parse all used-range cells from an XLSX workbook."""
    value_workbook = openpyxl.load_workbook(BytesIO(content), data_only=True)
    formula_workbook = openpyxl.load_workbook(BytesIO(content), data_only=False)
    cells: list[SourceCell] = []
    for sheet in value_workbook.worksheets:
        formula_sheet = formula_workbook[sheet.title]
        for row_index in range(1, sheet.max_row + 1):
            for column_index in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_index, column=column_index)
                formula_cell = formula_sheet.cell(
                    row=row_index,
                    column=column_index,
                )
                raw_value = _xlsx_cell_raw_value(cell.value)
                formula = (
                    str(formula_cell.value) if formula_cell.data_type == "f" else None
                )
                cells.append(
                    SourceCell(
                        artifact=artifact,
                        sheet_name=sheet.title,
                        row_number=row_index,
                        column_number=column_index,
                        address=cell.coordinate,
                        cell_type=_xlsx_cell_type(raw_value),
                        raw_value=raw_value,
                        display_value=None if raw_value is None else str(raw_value),
                        formula=formula,
                    )
                )
    return cells


def source_cells_from_ods(
    content: bytes,
    artifact: SourceArtifactMetadata,
) -> list[SourceCell]:
    """Parse all used-range cells from an ODS workbook."""
    with ZipFile(BytesIO(content)) as archive:
        root = ElementTree.fromstring(archive.read("content.xml"))

    spreadsheet = root.find(".//office:spreadsheet", _ODS_NAMESPACES)
    if spreadsheet is None:
        return []

    cells: list[SourceCell] = []
    for table in spreadsheet.findall("table:table", _ODS_NAMESPACES):
        sheet_name = table.attrib.get(_ods_attr("table", "name"), "Sheet")
        rows = _ods_rows(table)
        max_column = max((len(row) for row in rows), default=0)
        for row_index, row in enumerate(rows, start=1):
            for column_index in range(1, max_column + 1):
                raw_value = row[column_index - 1] if column_index <= len(row) else None
                cells.append(
                    SourceCell(
                        artifact=artifact,
                        sheet_name=sheet_name,
                        row_number=row_index,
                        column_number=column_index,
                        address=f"{_excel_column_name(column_index)}{row_index}",
                        cell_type=_scalar_cell_type(raw_value),
                        raw_value=raw_value,
                        display_value=None if raw_value is None else str(raw_value),
                    )
                )
    return cells


def source_cells_from_html_tables_and_text(
    content: bytes,
    artifact: SourceArtifactMetadata,
) -> list[SourceCell]:
    """Parse HTML tables and numeric document text into source cells."""
    parser = _HtmlTableAndTextParser()
    parser.feed(content.decode("utf-8", errors="replace"))
    parser.close()

    cells: list[SourceCell] = []
    for table_index, rows in enumerate(parser.tables, start=1):
        cells.extend(
            _html_table_cells(
                rows,
                artifact,
                sheet_name=f"table_{table_index}",
            )
        )
    cells.extend(_html_document_number_cells(parser.blocks, artifact))
    return cells


def source_cells_from_pdf_text_numbers(
    content: bytes,
    artifact: SourceArtifactMetadata,
) -> list[SourceCell]:
    """Parse PDF text lines into numeric document source cells."""
    from pypdf import PdfReader

    reader = PdfReader(BytesIO(content))
    sheet_name = "document_numbers"
    headers: tuple[Scalar, ...] = (
        "page_number",
        "line_number",
        "text",
        "number_text",
        "numeric_value",
        "context_text",
    )
    cells = [
        _html_document_cell(
            artifact,
            sheet_name,
            row_number=1,
            column_number=index + 1,
            raw_value=value,
        )
        for index, value in enumerate(headers)
    ]
    row_number = 2
    for page_number, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        lines = [
            (line_number, _normalize_html_text(line))
            for line_number, line in enumerate(text.splitlines(), start=1)
            if _normalize_html_text(line)
        ]
        for line_index, (line_number, normalized_line) in enumerate(lines):
            context_text = " ".join(
                line_text
                for _, line_text in lines[
                    max(0, line_index - 2) : min(len(lines), line_index + 3)
                ]
            )
            for match in _HTML_NUMBER_RE.finditer(normalized_line):
                number_text = match.group(0)
                for column_number, raw_value in enumerate(
                    (
                        page_number,
                        line_number,
                        normalized_line,
                        number_text,
                        _html_number_scalar(number_text),
                        context_text,
                    ),
                    start=1,
                ):
                    cells.append(
                        _html_document_cell(
                            artifact,
                            sheet_name,
                            row_number=row_number,
                            column_number=column_number,
                            raw_value=raw_value,
                        )
                    )
                row_number += 1
    return cells


def source_cells_from_delimited_text(
    content: bytes,
    artifact: SourceArtifactMetadata,
    *,
    sheet_name: str,
    selected_rows: tuple[dict[str, str], ...] = (),
    delimiter: str = ",",
) -> list[SourceCell]:
    """Parse selected rows from a delimited text artifact into source cells.

    The emitted row numbers are package-local so compact source-record specs can
    stay stable even when the publisher file is large. Data cells carry the
    original source line number in ``note``.
    """
    text = content.decode("utf-8-sig")
    reader = csv.reader(StringIO(text), delimiter=delimiter)
    try:
        header = next(reader)
    except StopIteration:
        return []

    normalized_header = [_normalize_header_cell(item) for item in header]
    selected_by_index: dict[int, tuple[int, list[str]]] = {}
    selected_indices_by_key_tuple: dict[
        tuple[str, ...], dict[tuple[str, ...], list[int]]
    ] = {}
    if selected_rows:
        for selection_index, criteria in enumerate(selected_rows):
            key_tuple = tuple(criteria)
            value_tuple = tuple(criteria.values())
            selected_indices_by_key_tuple.setdefault(key_tuple, {}).setdefault(
                value_tuple, []
            ).append(selection_index)
    for source_line_number, row in enumerate(reader, start=2):
        row_by_header = {
            normalized_header[index]: row[index] if index < len(row) else ""
            for index in range(len(normalized_header))
        }
        if not selected_rows:
            selected_by_index[len(selected_by_index)] = (source_line_number, row)
            continue
        for key_tuple, indices_by_value_tuple in selected_indices_by_key_tuple.items():
            value_tuple = tuple(row_by_header.get(key, "") for key in key_tuple)
            for selection_index in indices_by_value_tuple.get(value_tuple, ()):
                if selection_index not in selected_by_index:
                    selected_by_index[selection_index] = (source_line_number, row)

    cells = [
        _delimited_cell(
            artifact,
            sheet_name,
            row_number=1,
            column_number=index + 1,
            raw_value=value,
        )
        for index, value in enumerate(header)
    ]
    for virtual_row_number, selection_index in enumerate(
        sorted(selected_by_index),
        start=2,
    ):
        source_line_number, row = selected_by_index[selection_index]
        for column_index, value in enumerate(row, start=1):
            cells.append(
                _delimited_cell(
                    artifact,
                    sheet_name,
                    row_number=virtual_row_number,
                    column_number=column_index,
                    raw_value=_delimited_scalar(value),
                    note=f"source_line_number={source_line_number}",
                )
            )
    return cells


def validate_source_cells(cells: list[SourceCell]) -> SourceCellReport:
    """Validate source cells and return QA counts plus issues."""
    errors: list[SourceCellIssue] = []
    key_indices: dict[str, list[int]] = {}

    for index, cell in enumerate(cells):
        key = build_source_cell_key(cell)
        key_indices.setdefault(key, []).append(index)
        if not cell.artifact.source_name.strip():
            errors.append(
                SourceCellIssue(
                    code="missing_source_name",
                    message="Cell artifact is missing source_name",
                    source_cell_key=key,
                    cell_index=index,
                )
            )
        if not cell.artifact.sha256.strip():
            errors.append(
                SourceCellIssue(
                    code="missing_artifact_sha256",
                    message="Cell artifact is missing sha256",
                    source_cell_key=key,
                    cell_index=index,
                )
            )
        if cell.row_number < 1 or cell.column_number < 1:
            errors.append(
                SourceCellIssue(
                    code="malformed_coordinate",
                    message="Cell row and column numbers must be one-based",
                    source_cell_key=key,
                    cell_index=index,
                )
            )

    for key, indices in key_indices.items():
        if len(indices) > 1:
            errors.append(
                SourceCellIssue(
                    code="duplicate_source_cell_key",
                    message=f"Duplicate source-cell key appears at indices {indices}",
                    source_cell_key=key,
                    cell_index=indices[0],
                )
            )

    return SourceCellReport(
        cell_count=len(cells),
        counts=source_cell_counts(cells),
        errors=tuple(errors),
    )


def source_cell_counts(cells: list[SourceCell]) -> dict[str, dict[str, int]]:
    """Count cells across QA dimensions."""
    return {
        "by_source": _counter_dict(cell.artifact.source_name for cell in cells),
        "by_sheet": _counter_dict(cell.sheet_name for cell in cells),
        "by_cell_type": _counter_dict(cell.cell_type for cell in cells),
        "non_empty": {
            "count": sum(
                1 for cell in cells if cell.cell_type not in {"empty", "blank"}
            )
        },
    }


def load_source_cells_jsonl(path: str | Path) -> list[SourceCell]:
    """Load source cells from JSON Lines."""
    cell_path = Path(path)
    cells: list[SourceCell] = []
    with cell_path.open() as file:
        for line_number, line in enumerate(file, start=1):
            if not line.strip():
                continue
            try:
                payload = json.loads(line)
            except json.JSONDecodeError as exc:
                raise ValueError(
                    f"Invalid JSON on line {line_number} of {cell_path}"
                ) from exc
            cells.append(source_cell_from_mapping(payload))
    return cells


def save_source_cells_jsonl(cells: list[SourceCell], path: str | Path) -> None:
    """Save source cells to JSON Lines."""
    cell_path = Path(path)
    cell_path.parent.mkdir(parents=True, exist_ok=True)
    with cell_path.open("w") as file:
        for cell in cells:
            file.write(json.dumps(source_cell_to_mapping(cell), sort_keys=True))
            file.write("\n")


def source_cell_from_mapping(payload: dict[str, Any]) -> SourceCell:
    """Build a source cell from a JSON-compatible mapping."""
    return SourceCell(
        artifact=SourceArtifactMetadata(**payload["artifact"]),
        sheet_name=payload["sheet_name"],
        row_number=payload["row_number"],
        column_number=payload["column_number"],
        address=payload["address"],
        cell_type=payload["cell_type"],
        raw_value=payload["raw_value"],
        display_value=payload["display_value"],
        formula=payload.get("formula"),
        note=payload.get("note"),
        source_row_key=payload.get("source_row_key"),
    )


def source_cell_to_mapping(cell: SourceCell) -> dict[str, Any]:
    """Convert a source cell to a JSON-compatible mapping."""
    return asdict(cell)


def _cell_raw_value(cell: xlrd.sheet.Cell, datemode: int) -> Scalar:
    if cell.ctype in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK}:
        return None
    if cell.ctype == xlrd.XL_CELL_TEXT:
        return str(cell.value)
    if cell.ctype == xlrd.XL_CELL_NUMBER:
        value = float(cell.value)
        return int(value) if value.is_integer() else value
    if cell.ctype == xlrd.XL_CELL_DATE:
        return datetime(*xlrd.xldate_as_tuple(cell.value, datemode)).isoformat()
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    if cell.ctype == xlrd.XL_CELL_ERROR:
        return str(cell.value)
    return str(cell.value)


def _xlsx_cell_raw_value(value: Any) -> Scalar:
    if value is None:
        return None
    if isinstance(value, bool | int | float | str):
        return value
    if isinstance(value, datetime | date | time):
        return value.isoformat()
    return str(value)


def _xlsx_cell_type(raw_value: Scalar) -> str:
    if raw_value is None:
        return "empty"
    if isinstance(raw_value, bool):
        return "boolean"
    if isinstance(raw_value, int | float):
        return "number"
    return "text"


def _cell_type_name(cell_type: int) -> str:
    return {
        xlrd.XL_CELL_EMPTY: "empty",
        xlrd.XL_CELL_TEXT: "text",
        xlrd.XL_CELL_NUMBER: "number",
        xlrd.XL_CELL_DATE: "date",
        xlrd.XL_CELL_BOOLEAN: "boolean",
        xlrd.XL_CELL_ERROR: "error",
        xlrd.XL_CELL_BLANK: "blank",
    }.get(cell_type, "unknown")


_ODS_NAMESPACES = {
    "office": "urn:oasis:names:tc:opendocument:xmlns:office:1.0",
    "table": "urn:oasis:names:tc:opendocument:xmlns:table:1.0",
    "text": "urn:oasis:names:tc:opendocument:xmlns:text:1.0",
}


def _ods_attr(namespace: str, name: str) -> str:
    return f"{{{_ODS_NAMESPACES[namespace]}}}{name}"


def _ods_rows(table: ElementTree.Element) -> list[list[Scalar]]:
    rows: list[list[Scalar]] = []
    for row in table.findall("table:table-row", _ODS_NAMESPACES):
        row_values = _ods_row_values(row)
        row_repeat = int(row.attrib.get(_ods_attr("table", "number-rows-repeated"), 1))
        if not any(value is not None for value in row_values):
            row_repeat = min(row_repeat, 1)
        rows.extend([list(row_values) for _ in range(row_repeat)])

    while rows and not any(value is not None for value in rows[-1]):
        rows.pop()
    return rows


def _ods_row_values(row: ElementTree.Element) -> list[Scalar]:
    values: list[Scalar] = []
    for cell in row.findall("table:table-cell", _ODS_NAMESPACES):
        raw_value = _ods_cell_raw_value(cell)
        column_repeat = int(
            cell.attrib.get(_ods_attr("table", "number-columns-repeated"), 1)
        )
        if raw_value is None:
            column_repeat = min(column_repeat, 2048)
        values.extend([raw_value] * column_repeat)

    while values and values[-1] is None:
        values.pop()
    return values


def _ods_cell_raw_value(cell: ElementTree.Element) -> Scalar:
    value_type = cell.attrib.get(_ods_attr("office", "value-type"))
    if value_type in {"float", "currency", "percentage"}:
        value = cell.attrib.get(_ods_attr("office", "value"))
        if value is not None:
            numeric = float(value)
            return int(numeric) if numeric.is_integer() else numeric
    if value_type == "boolean":
        value = cell.attrib.get(_ods_attr("office", "boolean-value"))
        if value is not None:
            return value == "true"
    if value_type in {"date", "time"}:
        return cell.attrib.get(_ods_attr("office", f"{value_type}-value"))

    text = _ods_cell_text(cell)
    return text or None


def _ods_cell_text(cell: ElementTree.Element) -> str:
    paragraphs = []
    for paragraph in cell.findall("text:p", _ODS_NAMESPACES):
        text = "".join(paragraph.itertext()).strip()
        if text:
            paragraphs.append(" ".join(text.split()))
    return " ".join(paragraphs).strip()


_HTML_BLOCK_TAGS = {"h1", "h2", "h3", "h4", "h5", "h6", "p", "li"}
_HTML_NUMBER_RE = re.compile(
    r"(?<![\w.])[£$€]?-?(?:\d{1,3}(?:,\d{3})+|\d+)(?:\.\d+)?"
    r"(?:\s*(?:thousand|million|billion)|bn)?(?![\w.])",
    flags=re.IGNORECASE,
)


@dataclass(frozen=True)
class _HtmlCellValue:
    raw_value: Scalar
    display_value: str | None


@dataclass(frozen=True)
class _HtmlBlock:
    element_index: int
    tag: str
    text: str


class _HtmlTableAndTextParser(HTMLParser):
    """Small HTML parser for source preservation, not browser rendering."""

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.tables: list[list[list[_HtmlCellValue]]] = []
        self.blocks: list[_HtmlBlock] = []
        self._table_depth = 0
        self._current_table: list[list[_HtmlCellValue]] | None = None
        self._current_row: list[_HtmlCellValue] | None = None
        self._current_cell_parts: list[str] | None = None
        self._current_cell_colspan = 1
        self._current_block_tag: str | None = None
        self._current_block_parts: list[str] = []
        self._next_block_index = 1

    def handle_starttag(
        self,
        tag: str,
        attrs: list[tuple[str, str | None]],
    ) -> None:
        tag = tag.lower()
        if tag == "table":
            self._table_depth += 1
            if self._table_depth == 1:
                self._current_table = []
            return

        if self._table_depth:
            self._handle_table_starttag(tag, attrs)
            return

        if tag in _HTML_BLOCK_TAGS and self._current_block_tag is None:
            self._current_block_tag = tag
            self._current_block_parts = []
        elif tag == "br" and self._current_block_tag is not None:
            self._current_block_parts.append(" ")

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if self._table_depth:
            self._handle_table_endtag(tag)
            if tag == "table":
                self._table_depth -= 1
                if self._table_depth == 0 and self._current_table is not None:
                    self.tables.append(_trim_html_rows(self._current_table))
                    self._current_table = None
            return

        if tag == self._current_block_tag:
            text = _normalize_html_text("".join(self._current_block_parts))
            if text:
                self.blocks.append(
                    _HtmlBlock(
                        element_index=self._next_block_index,
                        tag=tag,
                        text=text,
                    )
                )
                self._next_block_index += 1
            self._current_block_tag = None
            self._current_block_parts = []

    def handle_data(self, data: str) -> None:
        if self._current_cell_parts is not None:
            self._current_cell_parts.append(data)
            return
        if not self._table_depth and self._current_block_tag is not None:
            self._current_block_parts.append(data)

    def _handle_table_starttag(
        self,
        tag: str,
        attrs: list[tuple[str, str | None]],
    ) -> None:
        if tag == "tr" and self._table_depth == 1:
            self._current_row = []
            return
        if tag in {"td", "th"} and self._current_row is not None:
            self._current_cell_parts = []
            self._current_cell_colspan = _html_colspan(attrs)
        elif tag == "br" and self._current_cell_parts is not None:
            self._current_cell_parts.append(" ")

    def _handle_table_endtag(self, tag: str) -> None:
        if tag in {"td", "th"} and self._current_cell_parts is not None:
            text = _normalize_html_text("".join(self._current_cell_parts))
            value = _html_cell_value(text)
            if self._current_row is not None:
                self._current_row.extend(
                    value for _ in range(self._current_cell_colspan)
                )
            self._current_cell_parts = None
            self._current_cell_colspan = 1
            return
        if tag == "tr" and self._current_row is not None:
            if self._current_table is not None:
                self._current_table.append(self._current_row)
            self._current_row = None


def _html_table_cells(
    rows: list[list[_HtmlCellValue]],
    artifact: SourceArtifactMetadata,
    *,
    sheet_name: str,
) -> list[SourceCell]:
    max_column = max((len(row) for row in rows), default=0)
    cells = []
    for row_index, row in enumerate(rows, start=1):
        for column_index in range(1, max_column + 1):
            value = (
                row[column_index - 1]
                if column_index <= len(row)
                else _HtmlCellValue(None, None)
            )
            cells.append(
                SourceCell(
                    artifact=artifact,
                    sheet_name=sheet_name,
                    row_number=row_index,
                    column_number=column_index,
                    address=f"{_excel_column_name(column_index)}{row_index}",
                    cell_type=_scalar_cell_type(value.raw_value),
                    raw_value=value.raw_value,
                    display_value=value.display_value,
                )
            )
    return cells


def _html_document_number_cells(
    blocks: list[_HtmlBlock],
    artifact: SourceArtifactMetadata,
) -> list[SourceCell]:
    sheet_name = "document_numbers"
    headers: tuple[Scalar, ...] = (
        "element_index",
        "element_type",
        "text",
        "number_text",
        "numeric_value",
    )
    cells = [
        _html_document_cell(
            artifact,
            sheet_name,
            row_number=1,
            column_number=index + 1,
            raw_value=value,
        )
        for index, value in enumerate(headers)
    ]
    row_number = 2
    for block in blocks:
        for match in _HTML_NUMBER_RE.finditer(block.text):
            number_text = match.group(0)
            for column_number, raw_value in enumerate(
                (
                    block.element_index,
                    block.tag,
                    block.text,
                    number_text,
                    _html_number_scalar(number_text),
                ),
                start=1,
            ):
                cells.append(
                    _html_document_cell(
                        artifact,
                        sheet_name,
                        row_number=row_number,
                        column_number=column_number,
                        raw_value=raw_value,
                    )
                )
            row_number += 1
    return cells


def _html_document_cell(
    artifact: SourceArtifactMetadata,
    sheet_name: str,
    *,
    row_number: int,
    column_number: int,
    raw_value: Scalar,
) -> SourceCell:
    return SourceCell(
        artifact=artifact,
        sheet_name=sheet_name,
        row_number=row_number,
        column_number=column_number,
        address=f"{_excel_column_name(column_number)}{row_number}",
        cell_type=_scalar_cell_type(raw_value),
        raw_value=raw_value,
        display_value=None if raw_value is None else str(raw_value),
    )


def _html_colspan(attrs: list[tuple[str, str | None]]) -> int:
    attr_dict = {key.lower(): value for key, value in attrs}
    try:
        return max(int(attr_dict.get("colspan") or 1), 1)
    except ValueError:
        return 1


def _html_cell_value(text: str) -> _HtmlCellValue:
    return _HtmlCellValue(_html_scalar(text), text or None)


def _html_scalar(text: str) -> Scalar:
    if not text:
        return None
    numeric = text.replace(",", "")
    if numeric.lstrip("-").isdigit():
        return int(numeric)
    try:
        return float(numeric)
    except ValueError:
        return text


def _html_number_scalar(text: str) -> int | float:
    normalized = text.replace(",", "").lower().lstrip("£$€")
    multiplier = 1
    for suffix, value in (
        ("thousand", 1_000),
        ("million", 1_000_000),
        ("billion", 1_000_000_000),
        ("bn", 1_000_000_000),
    ):
        if normalized.endswith(suffix):
            multiplier = value
            normalized = normalized[: -len(suffix)].strip()
            break
    value = float(normalized)
    scaled = value * multiplier
    return int(scaled) if scaled.is_integer() else scaled


def _normalize_html_text(text: str) -> str:
    text = text.replace("\xa0", " ").replace("\u202f", " ")
    return " ".join(text.split()).strip()


def _trim_html_rows(rows: list[list[_HtmlCellValue]]) -> list[list[_HtmlCellValue]]:
    trimmed = [list(row) for row in rows]
    while trimmed and not any(cell.raw_value is not None for cell in trimmed[-1]):
        trimmed.pop()
    max_column = max((len(row) for row in trimmed), default=0)
    while max_column:
        if any(
            max_column <= len(row) and row[max_column - 1].raw_value is not None
            for row in trimmed
        ):
            break
        max_column -= 1
    return [row[:max_column] for row in trimmed]


def _excel_column_name(column_number: int) -> str:
    name = ""
    while column_number:
        column_number, remainder = divmod(column_number - 1, 26)
        name = f"{chr(65 + remainder)}{name}"
    return name


def _delimited_cell(
    artifact: SourceArtifactMetadata,
    sheet_name: str,
    *,
    row_number: int,
    column_number: int,
    raw_value: Scalar,
    note: str | None = None,
    source_row_key: str | None = None,
) -> SourceCell:
    return SourceCell(
        artifact=artifact,
        sheet_name=sheet_name,
        row_number=row_number,
        column_number=column_number,
        address=f"{_excel_column_name(column_number)}{row_number}",
        cell_type=_scalar_cell_type(raw_value),
        raw_value=raw_value,
        display_value=None if raw_value is None else str(raw_value),
        note=note,
        source_row_key=source_row_key,
    )


def _delimited_scalar(value: str) -> Scalar:
    stripped = value.strip()
    if not stripped:
        return None
    numeric = stripped.replace("$", "").replace(",", "")
    if numeric.lstrip("-").isdigit():
        return int(numeric)
    try:
        return float(numeric)
    except ValueError:
        return stripped


def _normalize_header_cell(value: str) -> str:
    return value.strip().removeprefix("%")


def _scalar_cell_type(value: Scalar) -> str:
    if value is None:
        return "empty"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, int | float):
        return "number"
    return "text"


def _counter_dict(values: Any) -> dict[str, int]:
    return dict(sorted(Counter(values).items()))
